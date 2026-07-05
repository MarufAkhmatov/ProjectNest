"""Multi-format Jira export parser: CSV / XLSX / HTML -> list[dict] of raw rows."""
import csv
import io
import re
import json
import html as _htmllib
from pathlib import Path


def parse_file(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    data = path.read_bytes()
    if suffix == ".csv":
        return _parse_csv(data)
    if suffix in (".xlsx", ".xlsm"):
        return _parse_xlsx(data)
    if suffix in (".html", ".htm"):
        return _parse_html(data)
    # try CSV as a fallback
    return _parse_csv(data)


def _decode(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _parse_csv(data: bytes) -> list[dict]:
    text = _decode(data)
    sample = text[:4096]
    delim = ","
    try:
        delim = csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
    except Exception:
        pass
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = list(reader)
    if not rows:
        return []
    header = [h.strip() for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(c.strip() for c in r):
            continue
        # Jira CSV can repeat columns (e.g. multiple "Status"/"Outward issue link");
        # collapse duplicates into a comma-joined string.
        rec: dict[str, str] = {}
        for i, col in enumerate(header):
            val = r[i].strip() if i < len(r) else ""
            if col in rec and val:
                rec[col] = f"{rec[col]}|{val}" if rec[col] else val
            else:
                rec.setdefault(col, val)
        out.append(rec)
    return out


def is_history_file(path: Path) -> bool:
    """Detect a Jira 'Export → History (Current fields)' export by its columns."""
    try:
        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            hdr = [str(c.value).strip().lower() if c.value else "" for c in next(wb.active.iter_rows(max_row=1))]
        else:
            text = _decode(path.read_bytes())[:8192]
            line = text.splitlines()[0] if text else ""
            hdr = [h.strip().strip('"').lower() for h in re.split(r"[,;\t]", line)]
        return ("changed field" in hdr) or ("change time" in hdr) or ("изменённое поле" in hdr)
    except Exception:
        return False


def _hist_rows(path: Path):
    """Yield header + value rows from a history export (XLSX or CSV)."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        it = wb.active.iter_rows(values_only=True)
        header = [str(h).strip().lower() if h is not None else "" for h in next(it)]
        for r in it:
            yield header, r
    else:
        for rec in _parse_csv(path.read_bytes()):
            header = [k.strip().lower() for k in rec.keys()]
            yield header, list(rec.values())


def parse_jira_history(path: Path) -> dict:
    """Parse a Jira History export into per-issue status-transition timelines.

    Returns {issue_key: [{status, entered, exited, days}]}, status canonicalized.
    Exact TTM (Discovery/Delivery), lead time and flow are computed from this.
    """
    import datetime as dt
    from .normalize import canon_status, parse_date

    def to_dt(v):
        if isinstance(v, dt.datetime):
            return v
        return parse_date(str(v)) if v else None

    raw, cur, header = {}, None, None
    col = {}
    for header, r in _hist_rows(path):
        if not col:
            def idx(*names):
                for n in names:
                    if n in header:
                        return header.index(n)
                return -1
            col = {
                "key": idx("код", "code", "issue key", "key", "ключ проблемы"),
                "created": idx("создано", "created"),
                "resolved": idx("дата резолюции", "resolved", "resolution date", "дата решения"),
                "status": idx("статус", "status_current") if "status_current" in header else idx("статус", "status"),
                "ctime": idx("change time", "время изменения"),
                "field": idx("changed field", "изменённое поле", "измененное поле", "поле"),
                "old": idx("old value", "старое значение"),
                "new": idx("new value", "новое значение"),
            }
        g = lambda k: (r[col[k]] if col[k] >= 0 and col[k] < len(r) else None)
        key = g("key")
        if key:
            cur = str(key).strip()
            raw.setdefault(cur, {"created": g("created"), "resolved": g("resolved"), "events": []})
        if cur is None:
            continue
        fld = g("field")
        if fld and str(fld).strip().lower() == "status":
            t = to_dt(g("ctime"))
            new_v = g("new")
            old_v = g("old")
            # Skip header-leak rows: a Jira export occasionally injects the literal
            # word "Statuses" into either the old- or new-value column (we saw 13
            # such events silently breaking TTM). Drop them rather than treat as
            # real statuses.
            if (str(new_v or "").strip().lower() == "statuses"
                    or str(old_v or "").strip().lower() == "statuses"):
                continue
            if t:
                raw[cur]["events"].append((t, old_v, new_v))

    out = {}
    for key, info in raw.items():
        evs = sorted([e for e in info["events"] if isinstance(e[0], dt.datetime)], key=lambda x: x[0])
        if not evs:
            continue
        created = to_dt(info["created"]) or evs[0][0]
        resolved = to_dt(info["resolved"])
        segs, prev_t, prev_s = [], created, canon_status(str(evs[0][1] or ""))
        for (t, _old, new) in evs:
            d = max(0.0, (t - prev_t).total_seconds() / 86400.0)
            # Drop DEAD-status segments (canon_status -> "") so they never enter
            # TTM/lead-time math. The time spent "in" a dead step disappears from
            # phase totals — by user request these are not real process phases.
            if prev_s:
                segs.append({"status": prev_s, "entered": prev_t.isoformat(),
                             "exited": t.isoformat(), "days": round(d, 3)})
            prev_t, prev_s = t, canon_status(str(new or ""))
        end = resolved or evs[-1][0]
        d = max(0.0, (end - prev_t).total_seconds() / 86400.0)
        if not prev_s:
            # Last segment is dead — don't emit it.
            out[key] = segs
            continue
        segs.append({"status": prev_s, "entered": prev_t.isoformat(),
                     "exited": end.isoformat(), "days": round(d, 3)})
        out[key] = segs
    return out


def _parse_xlsx(data: bytes) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if r is None or not any(c is not None and str(c).strip() for c in r):
            continue
        rec = {}
        for i, col in enumerate(header):
            v = r[i] if i < len(r) else None
            rec[col] = "" if v is None else str(v).strip()
        out.append(rec)
    return out


def _parse_html(data: bytes) -> list[dict]:
    text = _decode(data)
    # Jira "Printable" / Word-HTML detail export: each issue is a card with
    # [KEY] <a href=.../browse/KEY> and a label/value grid (often RU/UZ locale).
    if re.search(r"\[[A-Z]{2,}-\d+\]\s*(&nbsp;)?\s*<a\s+href=", text) or 'class="formtitle"' in text:
        rows = _parse_jira_printable(text)
        if rows:
            return rows
    # Jira issue-navigator "Excel HTML" export: one big <table id="issuetable">
    # with per-column <td class="issuekey|summary|status|…"> cells. Identify cells
    # by their stable CSS class so locale-translated headers don't matter.
    if 'id="issuetable"' in text or 'class="issuetable"' in text or "headerrow-issuekey" in text:
        rows = _parse_jira_issuetable(text)
        if rows:
            return rows
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(text, "html.parser")
    except Exception:
        return _parse_html_stdlib(text)
    table = soup.find("table")
    if not table:
        return []
    rows = table.find_all("tr")
    header_cells = rows[0].find_all(["th", "td"])
    header = [c.get_text(strip=True) for c in header_cells]
    out = []
    for tr in rows[1:]:
        cells = tr.find_all(["td", "th"])
        if not cells:
            continue
        rec = {}
        for i, col in enumerate(header):
            rec[col] = cells[i].get_text(strip=True) if i < len(cells) else ""
        if any(rec.values()):
            out.append(rec)
    return out


# ---- Jira "Printable" detail HTML parser -----------------------------------
def _clean(s: str) -> str:
    s = re.sub(r"<[^>]+>", " ", s or "")
    s = _htmllib.unescape(s)
    return re.sub(r"\s+", " ", s).strip()


def _cell_raw(region: str, label: str) -> str:
    m = re.search(r"<b>\s*" + re.escape(label) + r"\s*:?\s*</b>\s*</td>\s*<td[^>]*>(.*?)</td>",
                  region, re.S)
    return m.group(1) if m else ""


def _parse_jira_printable(text: str) -> list[dict]:
    parts = re.split(r"(?=\[[A-Z]{2,}-\d+\]\s*(?:&nbsp;)?\s*<a\s+href=)", text)
    rows: list[dict] = []
    for p in parts:
        km = re.search(r"\[([A-Z]{2,}-\d+)\]", p)
        if not km:
            continue
        key = km.group(1)
        sm = re.search(r"/browse/" + re.escape(key) + r'"[^>]*>(.*?)</a>', p, re.S)
        summary = _clean(sm.group(1)) if sm else ""
        hm = re.search(r'href="([^"]*?/browse/' + re.escape(key) + r')"', p)
        url = hm.group(1) if hm else ""

        sub = re.search(r'subText[^>]*>(.*?)</span>', p, re.S)
        subtext = _clean(sub.group(1)) if sub else ""

        def dlabel(lbl: str) -> str:
            m = re.search(lbl + r"\s*:?\s*([0-3]?\d\.[01]?\d\.\d{4})", subtext)
            return m.group(1) if m else ""

        # label -> text value map from the grid
        fields: dict[str, str] = {}
        for lm in re.finditer(r"<b>\s*([^<:]+?)\s*:?\s*</b>\s*</td>\s*<td[^>]*>(.*?)</td>", p, re.S):
            lbl = _clean(lm.group(1))
            if lbl and lbl not in fields:
                fields[lbl] = _clean(lm.group(2))

        # comments (comment-header / comment-body pairs)
        comments = []
        for cm in re.finditer(
            r'<tr id="comment-header-(\d+)"[^>]*>(.*?)</tr>\s*<tr id="comment-body-\1"[^>]*>(.*?)</tr>',
            p, re.S,
        ):
            hdr, body = cm.group(2), cm.group(3)
            am = re.search(r'class="user-hover"[^>]*>(.*?)</a>', hdr)
            dm = re.search(r'#336699">\s*([\d.]+)\s*</font>', hdr)
            text = _clean(body)[:700]
            if text:
                comments.append({"author": _clean(am.group(1)) if am else "",
                                 "date": dm.group(1) if dm else "", "text": text})
        comments = comments[-12:]

        epic_raw = _cell_raw(p, "Epic Link")
        epic_key = ""
        em = re.search(r"/browse/([A-Z]{2,}-\d+)", epic_raw)
        if em:
            epic_key = em.group(1)
        blocks = "|".join(re.findall(r"/browse/([A-Z]{2,}-\d+)", _cell_raw(p, "Blocks")))

        rows.append({
            "Issue key": key,
            "url": url,
            "Summary": summary,
            "Status": fields.get("Статус") or fields.get("Status") or "",
            "Issue Type": fields.get("Тип") or fields.get("Type") or "",
            "PM": fields.get("PM") or "",
            "Assignee": fields.get("Исполнитель") or fields.get("Assignee") or "",
            "Reporter": fields.get("Автор") or fields.get("Reporter") or "",
            "Resolution": fields.get("Решение") or fields.get("Resolution") or "",
            "Project": key.split("-")[0],
            "Created": dlabel("Создано") or dlabel("Created"),
            "Updated": dlabel("Обновлен") or dlabel("Updated"),
            "Due Date": dlabel("Срок исполнения") or dlabel("Due"),
            "Epic Link": epic_key,
            "Blocks": blocks,
            "Приоритет": fields.get("Приоритет") or fields.get("Priority") or "",
            "Тип проекта": fields.get("Тип проекта") or "",
            "Требование регулятора": fields.get("Требование регулятора") or "",
            "Подразделение заказчика": fields.get("Подразделение заказчика") or "",
            "Скоринг-балл": fields.get("Скоринг-балл") or "",
            "Квартальный статус": fields.get("Квартальный статус") or "",
            # Owner (владелец / ФИО владельца) — business owner accountable for the item.
            "ФИО владельца": (fields.get("ФИО владельца") or fields.get("Владелец")
                              or fields.get("Owner") or ""),
            "comments_json": json.dumps(comments, ensure_ascii=False),
        })
    return rows


# Jira issuetable <td class="X"> → canonical column name normalize understands.
_IT_CLASS = {
    "issuekey": "Issue key", "summary": "Summary", "status": "Status",
    "issuetype": "Issue Type", "priority": "Priority", "resolution": "Resolution",
    "assignee": "Assignee", "reporter": "Reporter", "creator": "Creator",
    "created": "Created", "updated": "Updated", "resolutiondate": "Resolution Date",
    "duedate": "Due Date", "description": "Description", "project": "Project",
}


def _parse_jira_issuetable(text: str) -> list[dict]:
    """Parse the Jira issue-navigator Excel-HTML <table id="issuetable">.
    Rows are keyed by canonical English column names derived from each cell's
    CSS class (issuekey/summary/status/…), so the locale of the visible
    headers is irrelevant and nested tables inside a cell can't shift columns."""
    from html.parser import HTMLParser

    class IT(HTMLParser):
        def __init__(self):
            super().__init__()
            self.rows: list[dict] = []
            self.cur: dict | None = None
            self.cell = ""
            self.cell_field: str | None = None
            self.in_cell = False

        def handle_starttag(self, tag, attrs):
            if tag == "tr":
                self.cur = {}
            elif tag in ("td", "th"):
                self.in_cell = True
                self.cell = ""
                cls = dict(attrs).get("class", "") or ""
                self.cell_field = next((_IT_CLASS[t] for t in cls.split() if t in _IT_CLASS), None)

        def handle_endtag(self, tag):
            if tag in ("td", "th"):
                if self.cur is not None and self.cell_field and self.cell_field not in self.cur:
                    val = _htmllib.unescape(re.sub(r"\s+", " ", self.cell)).strip()
                    self.cur[self.cell_field] = val
                self.in_cell = False
                self.cell_field = None
            elif tag == "tr":
                # keep only real issue rows (those carrying an issue key)
                if self.cur and self.cur.get("Issue key"):
                    self.rows.append(self.cur)
                self.cur = None

        def handle_data(self, data):
            if self.in_cell:
                self.cell += data

    p = IT()
    p.feed(text)
    out = []
    for r in p.rows:
        # split a "Resolution Date" into the canonical "Resolved" column too
        if r.get("Resolution Date") and "Resolved" not in r:
            r["Resolved"] = r["Resolution Date"]
        out.append(r)
    return out


def _parse_html_stdlib(text: str) -> list[dict]:
    from html.parser import HTMLParser

    class T(HTMLParser):
        def __init__(self):
            super().__init__()
            self.rows, self.cur, self.cell, self.in_cell = [], None, "", False

        def handle_starttag(self, tag, attrs):
            if tag == "tr":
                self.cur = []
            elif tag in ("td", "th"):
                self.in_cell, self.cell = True, ""

        def handle_endtag(self, tag):
            if tag in ("td", "th") and self.cur is not None:
                self.cur.append(self.cell.strip())
                self.in_cell = False
            elif tag == "tr" and self.cur is not None:
                self.rows.append(self.cur)
                self.cur = None

        def handle_data(self, data):
            if self.in_cell:
                self.cell += data

    p = T()
    p.feed(text)
    if not p.rows:
        return []
    header = p.rows[0]
    out = []
    for r in p.rows[1:]:
        rec = {header[i]: (r[i] if i < len(r) else "") for i in range(len(header))}
        if any(rec.values()):
            out.append(rec)
    return out
